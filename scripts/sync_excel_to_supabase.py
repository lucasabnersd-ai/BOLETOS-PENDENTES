from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SUPABASE_URL = os.environ.get("BOLETOS_SUPABASE_URL", "https://pyrniqluywejmgzqkari.supabase.co").rstrip("/")
SUPABASE_KEY = os.environ.get("BOLETOS_SUPABASE_KEY", "sb_publishable_fXWQGDirOvs5xfxZDaSOtg_Jgd7vcbu")
ITEMS_TABLE = "boleto_pendentes_items"
META_TABLE = "boleto_pendentes_meta"
AUDIT_TABLE = "boleto_pendentes_audit"
DEFAULT_WORKBOOK = r"C:\Users\lucas\Grupo S&D\Gabriella Karla Oliveira Milas - FINANCEIRO COMPARTILHADO\LUCAS ABNER ARAUJO\BOLETOS PENDENTES A ASSOCIAR.xlsx"


FIELD_MAP = {
    "alerta": "alerta",
    "prioridade": "prioridade",
    "situacao_do_vencimento": "situacao_vencimento",
    "dias_para_vencer": "dias_para_vencer",
    "situacao_da_associacao": "situacao_associacao",
    "checklist": "checklist",
    "tratado_pendente": "tratado_pendente",
    "dda_itau": "dda_itau",
    "fonte": "fonte",
    "fornecedor": "fornecedor",
    "cnpj_cpf": "cnpj_cpf",
    "nf_doc_extraido": "nf_doc_extraido",
    "valor_rs": "valor",
    "vencimento": "vencimento",
    "data_emissao": "data_emissao",
    "linha_digitavel": "linha_digitavel",
    "codigo_de_barras_44": "codigo_barras",
    "origem": "origem",
    "remetente": "remetente",
    "assunto": "assunto",
    "observacao": "observacao",
}


def main() -> int:
    parser = argparse.ArgumentParser(description="Sincroniza BOLETOS PENDENTES A ASSOCIAR.xlsx com Supabase e GitHub Pages.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Caminho da planilha fonte.")
    parser.add_argument("--repo", default=str(Path(__file__).resolve().parents[1]), help="Pasta do repositorio.")
    parser.add_argument("--delete-missing", action="store_true", help="Remove da base online boletos que nao estao mais na planilha.")
    parser.add_argument("--no-delete-missing", dest="delete_missing", action="store_false")
    parser.set_defaults(delete_missing=True)
    args = parser.parse_args()

    workbook = Path(args.workbook)
    repo = Path(args.repo)
    if not workbook.exists():
      raise SystemExit(f"Planilha nao encontrada: {workbook}")

    rows, summary = read_workbook(workbook)
    if not rows:
      raise SystemExit("Nenhuma linha valida encontrada na aba Boletos Pendentes.")

    upsert_rows(rows)
    if args.delete_missing:
      delete_missing_rows({row["source_key"] for row in rows})

    meta = build_meta(workbook, rows, summary)
    upsert_meta(meta)
    insert_audit({"source": "sync_excel_to_supabase", "row_count": len(rows), "workbook": str(workbook)})
    write_snapshot(repo, rows, meta)
    print(f"OK: {len(rows)} boletos sincronizados em {dt.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}.")
    return 0


def read_workbook(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Boletos Pendentes"] if "Boletos Pendentes" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [header_key(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    duplicate_count = 0
    seen: set[str] = set()

    for row_index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
      raw = {headers[i]: clean_cell(values[i]) if i < len(values) else None for i in range(len(headers))}
      if not any(value not in (None, "") for value in raw.values()):
        continue

      item = {target: normalize_value(target, raw.get(source)) for source, target in FIELD_MAP.items()}
      if not item.get("fornecedor") and not item.get("linha_digitavel") and not item.get("codigo_barras"):
        continue

      item["source_key"] = make_source_key(item)
      if item["source_key"] in seen:
        duplicate_count += 1
        continue
      seen.add(item["source_key"])

      item["modelo"] = item.get("fonte") or infer_modelo(item)
      item["visao"] = item.get("origem") or item.get("fonte") or "Geral"
      item["raw"] = {source: to_jsonable(value) for source, value in raw.items() if source}
      item["raw"]["excel_row"] = row_index
      rows.append(item)

    total_value = sum(Decimal(str(row.get("valor") or "0")) for row in rows)
    summary = {
      "row_count": len(rows),
      "duplicate_count": duplicate_count,
      "total_value": float(total_value),
      "sources": count_by(rows, "fonte"),
      "origins": count_by(rows, "origem"),
      "priorities": count_by(rows, "prioridade"),
    }
    return rows, summary


def header_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    for old, new in [("/", "_"), ("(", ""), (")", ""), ("$", "s"), ("44", "44")]:
      text = text.replace(old, new)
    text = "".join(ch if ch.isalnum() else "_" for ch in text)
    while "__" in text:
      text = text.replace("__", "_")
    return text.strip("_")


def clean_cell(value: Any) -> Any:
    if isinstance(value, str):
      value = value.strip()
      return value if value else None
    return value


def normalize_value(field: str, value: Any) -> Any:
    if value in (None, ""):
      return None
    if field in {"vencimento", "data_emissao"}:
      return parse_date(value)
    if field == "valor":
      return parse_decimal(value)
    if field == "dias_para_vencer":
      try:
        return int(value)
      except (TypeError, ValueError):
        return None
    if field == "checklist":
      text = str(value).strip().upper()
      return text in {"1", "SIM", "OK", "TRUE", "VERDADEIRO", "X"}
    return str(value).strip()


def parse_date(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
      return value.date().isoformat()
    if isinstance(value, dt.date):
      return value.isoformat()
    text = str(value).strip()
    if not text:
      return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
      try:
        return dt.datetime.strptime(text[:10], fmt).date().isoformat()
      except ValueError:
        pass
    return None


def parse_decimal(value: Any) -> float | None:
    if value in (None, ""):
      return None
    if isinstance(value, (int, float, Decimal)):
      return float(value)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
      text = text.replace(".", "").replace(",", ".")
    try:
      return float(Decimal(text))
    except (InvalidOperation, ValueError):
      return None


def make_source_key(item: dict[str, Any]) -> str:
    parts = [
      digits(item.get("codigo_barras")),
      digits(item.get("linha_digitavel")),
      normalize_key_text(item.get("fonte")),
      normalize_key_text(item.get("fornecedor")),
      digits(item.get("cnpj_cpf")),
      normalize_key_text(item.get("nf_doc_extraido")),
      money_key(item.get("valor")),
      str(item.get("vencimento") or ""),
    ]
    basis = "|".join(part for part in parts if part)
    return hashlib.sha256(basis.encode("utf-8")).hexdigest()[:24]


def infer_modelo(item: dict[str, Any]) -> str:
    origem = normalize_key_text(item.get("origem"))
    if "itau" in origem:
      return "BOLETOS_ITAU"
    if "graziela" in origem:
      return "CENTRAL_GRAZIELA"
    if "lucas" in origem or "financeiro" in origem:
      return "CENTRAL_LUCAS"
    return "OUTROS"


def digits(value: Any) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def normalize_key_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return " ".join(text.split())


def money_key(value: Any) -> str:
    if value in (None, ""):
      return ""
    return f"{Decimal(str(value)).quantize(Decimal('0.01'))}"


def to_jsonable(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date)):
      return value.isoformat()
    if isinstance(value, Decimal):
      return float(value)
    return value


def count_by(rows: list[dict[str, Any]], field: str) -> dict[str, int]:
    result: dict[str, int] = {}
    for row in rows:
      key = str(row.get(field) or "Sem valor")
      result[key] = result.get(key, 0) + 1
    return dict(sorted(result.items(), key=lambda item: (-item[1], item[0])))


def build_meta(path: Path, rows: list[dict[str, Any]], summary: dict[str, Any]) -> dict[str, Any]:
    stat = path.stat()
    return {
      "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
      "workbook_path": str(path),
      "workbook_name": path.name,
      "workbook_modified_at": dt.datetime.fromtimestamp(stat.st_mtime, dt.timezone.utc).isoformat(),
      "row_count": len(rows),
      "duplicate_count": summary["duplicate_count"],
      "total_value": summary["total_value"],
      "sources": summary["sources"],
      "origins": summary["origins"],
      "priorities": summary["priorities"],
    }


def upsert_rows(rows: list[dict[str, Any]]) -> None:
    for chunk in chunks(rows, 100):
      request_json(
        "POST",
        f"/rest/v1/{ITEMS_TABLE}?on_conflict=source_key",
        chunk,
        prefer="resolution=merge-duplicates,return=minimal",
      )


def upsert_meta(meta: dict[str, Any]) -> None:
    request_json(
      "POST",
      f"/rest/v1/{META_TABLE}?on_conflict=key",
      [{"key": "latest_import", "value": meta}],
      prefer="resolution=merge-duplicates,return=minimal",
    )


def insert_audit(payload: dict[str, Any]) -> None:
    try:
      request_json(
        "POST",
        f"/rest/v1/{AUDIT_TABLE}",
        [{"field_name": "sync_excel_import", "old_value": None, "new_value": {"action": "sync_excel_import", "payload": payload}}],
        prefer="return=minimal",
      )
    except RuntimeError as exc:
      print(f"Aviso: auditoria nao gravada: {exc}")


def delete_missing_rows(current_keys: set[str]) -> None:
    existing = request_json("GET", f"/rest/v1/{ITEMS_TABLE}?select=source_key", None)
    for item in existing or []:
      source_key = item.get("source_key")
      if source_key and source_key not in current_keys:
        encoded = urllib.parse.quote(str(source_key), safe="")
        request_json("DELETE", f"/rest/v1/{ITEMS_TABLE}?source_key=eq.{encoded}", None, prefer="return=minimal")


def request_json(method: str, path: str, payload: Any = None, prefer: str | None = None) -> Any:
    data = None
    headers = {
      "apikey": SUPABASE_KEY,
      "Authorization": f"Bearer {SUPABASE_KEY}",
      "Content-Type": "application/json",
    }
    if prefer:
      headers["Prefer"] = prefer
    if payload is not None:
      data = json.dumps(payload, ensure_ascii=False, default=to_jsonable).encode("utf-8")
    request = urllib.request.Request(f"{SUPABASE_URL}{path}", data=data, headers=headers, method=method)
    try:
      with urllib.request.urlopen(request, timeout=60) as response:
        body = response.read().decode("utf-8")
        return json.loads(body) if body else None
    except urllib.error.HTTPError as exc:
      body = exc.read().decode("utf-8", errors="replace")
      raise RuntimeError(f"{method} {path} falhou: HTTP {exc.code} {body}") from exc
    except urllib.error.URLError as exc:
      raise RuntimeError(f"{method} {path} falhou: {exc}") from exc


def chunks(values: list[dict[str, Any]], size: int):
    for index in range(0, len(values), size):
      yield values[index:index + size]


def write_snapshot(repo: Path, rows: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    data_dir = repo / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    (data_dir / "initial.json").write_text(
      json.dumps({"meta": meta, "rows": rows}, ensure_ascii=False, indent=2, default=to_jsonable),
      encoding="utf-8",
    )
    (repo / "update-meta.json").write_text(
      json.dumps(meta, ensure_ascii=False, indent=2, default=to_jsonable),
      encoding="utf-8",
    )


if __name__ == "__main__":
    try:
      raise SystemExit(main())
    except Exception as exc:
      print(f"ERRO: {exc}", file=sys.stderr)
      raise
